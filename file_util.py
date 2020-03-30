from openpyxl import load_workbook


def get_excel_sheet():
    """
    Retrieves excel sheet from our vocabulary file. At the moment we know that we only have 1 sheet,
    so we can just retrieve the first one.
    :return: Excel sheet
    """
    workbook = load_workbook('vocab.xlsx', data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    return sheet


def get_entries(sheet):
    """
    Retrieves vocabulary entries
    :param sheet: Excel sheet from vocabulary file
    :return: Vocabulary entries as dictionary
    """
    vocabulary = {"entries": []}
    for row in sheet.iter_rows():
        entry = {"kanji": row[0].value,
                 "hiragana": row[1].value,
                 "english": row[2].value,
                 "level": row[3].value,
                 "examples": row[4].value}
        vocabulary["entries"].append(entry)
    return vocabulary
